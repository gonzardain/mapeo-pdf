# -*- coding: utf-8 -*-
######################################################################################################################################
#		Nombre Proyecto: CYF PDF MAKER
#		Autor: Gonzalo Zardain Castelazo
#		Empresa: Codelab (Grupo Teklan S. de R.L. de C.V.)
#       Fecha de inicio:
#       Fecha de termino
#       Notas:
#
#
#
#
#
#
#
#
#####################################################################################################################################

##################################################
#		Seccion de librerias a importar          #
##################################################

from pyPdf import PdfFileWriter, PdfFileReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm, inch
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code128, code39
from reportlab.platypus import Paragraph, Frame, KeepInFrame
import cStringIO
from openpyxl import load_workbook
from reportlab.lib.enums import TA_RIGHT
import math, os

##################################################
#		Ubicacion de las fuentes                 #
##################################################

base = os.path.dirname(os.path.abspath(__file__))
fonts_path = os.path.join(base, "fonts/")

KeepCalmTTF = "KeepCalm-Medium.ttf"
AvenirRegularTTF = "AvenirNextLTPro-Regular.ttf"
AvenirHeavyTTF = "AvenirNextLTPro-HeavyCn.ttf"
AvenirBoldTTF = "AvenirNextLTPro-Bold.ttf"
AvenirItaTTF = "AvenirNextLTPro-It.ttf"

font_KeepCalm = fonts_path + KeepCalmTTF
pdfmetrics.registerFont(TTFont("Keep Calm", font_KeepCalm))

font_AvenirRegular = fonts_path + AvenirRegularTTF
pdfmetrics.registerFont(TTFont("Avenir Regular", font_AvenirRegular))

font_AvenirHeavy = fonts_path + AvenirHeavyTTF
pdfmetrics.registerFont(TTFont("Avenir Heavy", font_AvenirHeavy))

font_AvenirBold = fonts_path + AvenirBoldTTF
pdfmetrics.registerFont(TTFont("Avenir Bold", font_AvenirBold))

font_AvenirItalic = fonts_path + AvenirItaTTF
pdfmetrics.registerFont(TTFont("Avenir Italic", font_AvenirItalic))


def load_wb():
	data_dict = []
	wb = load_workbook('data.xlsx', data_only=False)
	sheet_ranges = wb[wb.get_sheet_names()[0]]
	ws = wb.active
	row_count = ws.max_row
	column_count = ws.max_column
	for row in range(2, row_count+1):
		for column in range (1, column_count+1):
			data = ws.cell(column=column, row=row)
			data = data.value
			create_array(data, column, data_dict)

def create_array(data, column, data_dict):
	data_dict.append(data)
	if column == 34:
		do_stuff(data_dict)


def do_stuff(data_dict):
    #print data_dict
	stringBuffer = cStringIO.StringIO()
	writePage1(stringBuffer, data_dict)
	del data_dict[:]


def writePage1(stringBuffer, data_dict):
	settingBarcode(stringBuffer, data_dict)


def settingBarcode(stringBuffer, data_dict):
	barcode_value = "SPM290123456789123456"
	can = canvas.Canvas(stringBuffer, pagesize=letter)
	#barcode1 = code128.Code128("CA0390029",barWidth=.25*mm,barHeight=5*mm, checksum = True)
	barcode1 = code39.Standard39("SPM290123456789123456", stop=1, barHeight=15*mm, barWidth=.215*mm,  checksum=1)
	#barcode2 = code128.Code128("000715u057123",barWidht=.25*cm,barHeight =5*mm)
	settingText(stringBuffer, can, barcode1,  data_dict)

def settingText(stringBuffer, can, barcode1,  data_dict):
	global first_name, name
	count = 0
	coordinates_x = [
					425, #0 Cuenta#
					110, #1 Nombre#
					185, #2 Propietario#
					185, #3 Calle#
					185, #4 numero ext#
					185, #5 numero int#
					185, #6 colonia#
					185, #7 codigo postal#
					185, #8 municipio#
					185,  #9 estado#
					-72, #10 impuesto semestre 1#
					-72, #11 impuesto semestre 2#
					-72, #12 rezago#
					-72, #13 recargo#
					-72, #14 gto ejecucuion#
					-72, #15 multa#
					-72, #16 actualizacion#
					-72, #17 adeudo predial#
					176.206, #18 nada#
					-72, #19 consumo#
					-72, #20 alcantarillado#
					-72, #21 rezago#
					-72, #22 recargo#
					-72, #23 actualizacion#
					-72, #24 otros cargos#
					-72, #25 adeudo agua#
					70.823, #26 nada#
					70.823, #27 nada#
					70.823, #28 nada#
					-72, #29 total final#
					70.823, #30 ignorar#
					116.937, #31 captura predial#
					116.937, #32 captura agua#
					70.823, #33 ignorar#
					]

	coordinates_y = [
					875, #0 Cuenta#
					731.5, #1 Nombre#
					860, #2 Propietario#
					860, #3 Calle#
					860, #4 numero ext#
					860, #5 numero int#
					860, #6 colonia#
					860, #7 codigo postal#
					860, #8 municipio#
					860, #9 estado#
					415.79, #10 impuesto semestre 1#
					403.79, #11 impuesto semestre 2#
					391.79, #12 rezago#
					367.71, #13 recargo#
					355.79, #14 gto ejecucuion#
					343.79, #15 multa#
					379.79, #16 actualizacion#
					331.751, #17 adeudo predial#
					176.206, #18 nada#
					188.206, #19 consumo#
					176.206, #20 alcantarillado#
					164.206, #21 rezago#
					140.206, #22 recargo#
					152.206, #23 actualizacion#
					128.206, #24 otros cargos#
					116.162, #25 adeudo agua#
					70.823, #26 nada#
					70.823, #27 nada#
					70.823, #28 nada#
					70.823, #29 total final#
					70.823, #30 ignorar#
					431.198, #31 captura predial#
					431.198, #32 captura agua#
					70.823 #33 ignorar#
					]



	name = data_dict[2]
	first_name = data_dict[1]

	for data in range(10):
		if data_dict[data] is None:
			data_dict[data] = ""

	data_dict[4] = str(data_dict[4])
	data_dict[5] = str(data_dict[5])
	data_dict[7] = str(data_dict[7])

####################################################
#		Datos de Codigo de Barras                  #
####################################################

	text= '<font name= "Keep Calm" color= "#231F20" size="6">SPM%s0</font>'%data_dict[31]
	settingParagraphs6(text, can, stringBuffer, 400, 987, count, data_dict, barcode1)
	#print data_dict[23]

####################################################
#		Nombre y Direccion Frontal                 #
####################################################
	text= '<font name= "Keep Calm" color= "white" size="10">%s<br/><br/>%s %s %s %s CP %s <br/> %s %s</font>'%( data_dict[2], data_dict[3], data_dict[4][:-2], data_dict[5][:-2], data_dict[6], data_dict[7][:-2], data_dict[8], data_dict[9])
	settingParagraphs(text, can, stringBuffer, 185, 860, count, data_dict, barcode1)

####################################################
#		Numero de Cuenta Frontal                   #
####################################################

	text= '<font name= "Keep Calm" color= "white" size="8">CUENTA: %s</font>'%( str(data_dict[0]).zfill(6))
	settingParagraphs5(text, can, stringBuffer, 477, 875, count, data_dict, barcode1)

####################################################
#		Linea de Captura Predial                   #
####################################################

	text= '<font name= "Keep Calm" color= "#231F20" size="8">%s</font>'%data_dict[31]
	settingParagraphs5(text, can, stringBuffer, 116.937, 431.168, count, data_dict, barcode1)
	#print data_dict[23]

####################################################
#		Linea de Captura Agua                      #
####################################################

	text= '<font name= "Keep Calm" color= "#231F20" size="8">%s</font>'%data_dict[32]
	settingParagraphs5(text, can, stringBuffer, 116.937, 212.6, count, data_dict, barcode1)
	#print data_dict[23]

####################################################
#		Direccion en la seccion 2	               #
####################################################

	text2= '<font name= "Avenir Bold" color= "#4a4c4c" size="8">%s %s %s %s %s %s %s</font>'%( data_dict[3], data_dict[4][:-2], data_dict[5][:-2], data_dict[6], data_dict[7][:-2], data_dict[8], data_dict[9])
	settingParagraphs4(text2, can, stringBuffer, 40, 523.5, count, data_dict, barcode1)

	for index in range(len(data_dict)):
		text2 =""

####################################################
#		Datos predial                              #
####################################################
		if 9 < index < 17:
			value = float(data_dict[index])
			text2= '<font name= "Avenir Regular" Color="#231f20" size="9">${:,.2f}</font>'.format(value)

####################################################
#		Total Predial Redondeado                   #
####################################################
		elif index == 17:
			value = float(data_dict[index])
			value_rounded = math.ceil(value)
			text2= '<font name= "Avenir Bold" Color="#231f20"  size="9">${:,.2f}</font>'.format(value_rounded)

####################################################
#		Datos Agua                                 #
####################################################
		elif 18 < index < 25:
			value = float(data_dict[index])
			text2= '<font name= "Avenir Regular" size="9">${:,.2f}</font>'.format(value)

####################################################
#		Total Agua                                 #
####################################################
		elif index == 25:
			value = float(data_dict[index])
			value_rounded = math.ceil(value)
			text2= '<font name= "Avenir Bold" Color="#231f20"  size="9">${:,.2f}</font>'.format(value_rounded)
####################################################
#		Total a Pagar                              #
####################################################
		elif index == 29:
			value = float(data_dict[index])
			value_rounded = math.ceil(value)
			text2= '<font name= "Avenir Heavy" Color="#231f20"  size="10">${:,.2f}</font>'.format(value_rounded)

		count = count + 1
		settingParagraphs3(text2, can, stringBuffer, coordinates_x[count-1], coordinates_y[count-1], count, data_dict, barcode1)

def settingParagraphs6(text,  can, stringBuffer, x, y, count, data_dict, barcode1):

	style = getSampleStyleSheet()
	width, height = letter
	paragraph = Paragraph(text, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)
	if count == len(data_dict):
		saveCanvas(can, stringBuffer)

def settingParagraphs(text,  can, stringBuffer, x, y, count, data_dict, barcode1):
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, 285, 96, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	barcode1.drawOn(can, 325, 999, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
	if count == len(data_dict):
		saveCanvas(can, stringBuffer)


def settingParagraphs4(text,  can, stringBuffer, x, y, count, data_dict, barcode1):
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, 530, 40, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(7.2*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	if count == len(data_dict):
		saveCanvas(can, stringBuffer)

def settingParagraphs3(text2,  can, stringBuffer, x, y, count, data_dict, barcode1):
	style = getSampleStyleSheet()
	style.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT))
	width, height = letter
	paragraph = Paragraph(text2, style['RightAlign'])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)
	if count == len(data_dict):
		saveCanvas(can, stringBuffer)

def settingParagraphs5(text2,  can, stringBuffer, x, y, count, data_dict, barcode1):
	style = getSampleStyleSheet()
	style.add(ParagraphStyle(name='RightAlign', alignment=TA_RIGHT))
	width, height = letter
	paragraph = Paragraph(text2, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)
	if count == len(data_dict):
		saveCanvas(can, stringBuffer)


def saveCanvas(can, stringBuffer):
	can.save()
	createPDF(stringBuffer)


def createPDF(stringBuffer):
	global pdf_counter
	#print "createPDF"
	stringBuffer.seek(0)
	new_pdf = PdfFileReader(stringBuffer)
	existing_pdf = PdfFileReader(file("demo.pdf", "rb"))
	pages = existing_pdf.getNumPages()
	output = PdfFileWriter()
	page_1 = existing_pdf.getPage(0)
	page_2 = existing_pdf.getPage(1)
	page_1.mergePage(new_pdf.getPage(0))
	output.addPage(page_1)
	output.addPage(page_2)
	name = 'newpdf_%d.pdf' %pdf_counter
	#print "creating %s"%name
	outputStream = file(name, "wb")
	output.write(outputStream)
	outputStream.close()
	pdf_counter = pdf_counter + 1
	writePage2(cStringIO.StringIO())

def writePage2(stringBuffer):
	settingText2(stringBuffer)


def settingText2(stringBuffer):
	global first_name, name, c
	can = canvas.Canvas(stringBuffer, pagesize=letter)

	coordinates_x = [63]
	coordinates_y =[807]
	coordinates_xn = [110]
	coordinates_yn =[731]
	#font_AvenirBold = r"/home/gonz/cloud/Codelab/Software/cyf/fonts/AvenirNextLTPro-Bold.ttf"
	#pdfmetrics.registerFont(TTFont("Avenir Bold", font_AvenirBold))
	#font_AvenirItalic = r"/home/gonz/cloud/Codelab/Software/cyf/fonts/AvenirNextLTPro-It.ttf"
	#pdfmetrics.registerFont(TTFont("Avenir Italic", font_AvenirItalic))
	text= '<font name="Avenir Bold" color= "#4a4c4c" size="16">%s</font>'%name.title()
	settingParagraphs2(text, can, stringBuffer, coordinates_x[0], coordinates_y[0])
	text= '<font name="Avenir Italic" color= "#4a4c4c" size="12">%s</font>'%first_name.title()
	settingParagraphs2(text, can, stringBuffer, coordinates_xn[0], coordinates_yn[0])


def settingParagraphs2(text, can, stringBuffer, x, y ):
	global c
	c = c +1
	style = getSampleStyleSheet()
	width, height = letter
	paragraph = Paragraph(text, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)
	print c
	if c % 2 == 0:
		saveCanvas2(can, stringBuffer)

def saveCanvas2(can, stringBuffer):
	can.save()
	createPDF2(stringBuffer)


def createPDF2(stringBuffer):
	global first_name, pdf_counter
	cnt = pdf_counter -1
	stringBuffer.seek(0)
	new_pdf = PdfFileReader(stringBuffer)
	pdftowrite = "newpdf_%d.pdf"%cnt
	existing_pdf = PdfFileReader(file(pdftowrite, "rb"))
	pages = existing_pdf.getNumPages()
	output = PdfFileWriter()
	page_1 = existing_pdf.getPage(0)
	page_2 = existing_pdf.getPage(1)
	page_2.mergePage(new_pdf.getPage(0))
	output.addPage(page_1)
	output.addPage(page_2)
	name = 'pdf_%d.pdf' %cnt
	print 'pdf_%s.pdf' %first_name
	save_name = os.path.join(os.path.expanduser("~"), "Desktop/PDF", name)
	#print "creating %s from %s"%(name, first_name)
	outputStream = file(save_name, "wb")
	output.write(outputStream)
	outputStream.close()

if __name__ == "__main__":
	global pdf_counter, pdf_final, c
	c = 0
	pdf_final= 1
	pdf_counter = 1
	load_wb()
	#writePage1(cStringIO.StringIO())
	#writePage2(cStringIO.StringIO())
