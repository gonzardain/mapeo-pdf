from pyPdf import PdfFileWriter, PdfFileReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm, inch
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.barcode import code128, code39
from reportlab.platypus import Paragraph, Frame, KeepInFrame
import cStringIO
from openpyxl import load_workbook
from reportlab.lib.enums import TA_RIGHT
import math, os, checksum

##################################################
#		Ubicacion de las fuentes                 #
##################################################

base = os.path.dirname(os.path.abspath(__file__))
fonts_path = os.path.join(base, "fonts/")

KeepCalmTTF = "KeepCalm-Medium.ttf"
AvenirRegularTTF = "AvenirNextLTPro-Regular.ttf"
AvenirHeavyTTF = "AvenirNextLTPro-HeavyCn.ttf"
AvenirBoldTTF = "AvenirNextLTPro-Bold.ttf"
AvenirItaTTF = "AvenirNextLTPro-It.ttf"
TimesTTF = "times.ttf"

font_KeepCalm = fonts_path + KeepCalmTTF
pdfmetrics.registerFont(TTFont("Keep Calm", font_KeepCalm))

font_AvenirRegular = fonts_path + AvenirRegularTTF
pdfmetrics.registerFont(TTFont("Avenir Regular", font_AvenirRegular))

font_AvenirHeavy = fonts_path + AvenirHeavyTTF
pdfmetrics.registerFont(TTFont("Avenir Heavy", font_AvenirHeavy))

font_AvenirBold = fonts_path + AvenirBoldTTF
pdfmetrics.registerFont(TTFont("Avenir Bold", font_AvenirBold))

font_AvenirItalic = fonts_path + AvenirItaTTF
pdfmetrics.registerFont(TTFont("Avenir Italic", font_AvenirItalic))

font_Times = fonts_path + TimesTTF
pdfmetrics.registerFont(TTFont("Times", font_Times))

def load_wb():
	data_dict = []
	wb = load_workbook('data.xlsx', data_only=False)
	sheet_ranges = wb[wb.get_sheet_names()[0]]
	ws = wb.active
	row_count = ws.max_row
	column_count = ws.max_column
	for row in range(2, row_count+1):
		for column in range (1, column_count+1):
			data = ws.cell(column=column, row=row)
			data = data.value
			create_array(data, column, data_dict)

def create_array(data, column, data_dict):
	data_dict.append(data)
	if column == 36:
		do_stuff(data_dict)


def do_stuff(data_dict):
    #print data_dict
	stringBuffer = cStringIO.StringIO()
	writePage1(stringBuffer, data_dict)
	del data_dict[:]

def writePage1(stringBuffer, data_dict):
	settingBarcode(stringBuffer, data_dict)

def settingBarcode(stringBuffer, data_dict):
	value_default = "29"
	value_16 = str(data_dict[31].__add__("0"))
	total_value = "{}{}".format(value_default, value_16)
	checkdigit = str(checksum.add_check_digit(total_value))
	can = canvas.Canvas(stringBuffer, pagesize=letter)
	barcode_value = "CYF{}{}{}".format(value_default, value_16, checkdigit)
	barcode1 = code39.Standard39(barcode_value, stop=1, barHeight=15*mm, barWidth=.215*mm,  checksum=0)
	#barcode2 = code128.Code128("000715u057123",barWidht=.25*cm,barHeight =5*mm)
	settingText(stringBuffer, can, barcode1,  data_dict, barcode_value)

def settingText(stringBuffer, can, barcode1,  data_dict, barcode_value):
	global first_name, name
	count = 0
	coordinates_x = [
					425, #0 Cuenta#
					110, #1 Nombre#
					185, #2 Propietario#
					185, #3 Calle#
					185, #4 numero ext#
					185, #5 numero int#
					185, #6 colonia#
					185, #7 codigo postal#
					185, #8 municipio#
					185,  #9 estado#
					-72, #10 impuesto semestre 1#
					-72, #11 impuesto semestre 2#
					-72, #12 rezago#
					-72, #13 recargo#
					-72, #14 gto ejecucuion#
					-72, #15 multa#
					-72, #16 actualizacion#
					-72, #17 adeudo predial#
					176.206, #18 nada#
					-72, #19 consumo#
					-72, #20 alcantarillado#
					-72, #21 rezago#
					-72, #22 recargo#
					-72, #23 actualizacion#
					-72, #24 otros cargos#
					-72, #25 adeudo agua#
					70.823, #26 nada#
					70.823, #27 nada#
					70.823, #28 nada#
					-72, #29 total final#
					70.823, #30 ignorar#
					116.937, #31 captura predial#
					116.937, #32 captura agua#
					70.823, #33 adicional#
					70.823, #34 adicional#
					70.823, #35 adicional#

					]

	coordinates_y = [
					875, #0 Cuenta#
					731.5, #1 Nombre#
					860, #2 Propietario#
					860, #3 Calle#
					860, #4 numero ext#
					860, #5 numero int#
					860, #6 colonia#
					860, #7 codigo postal#
					860, #8 municipio#
					860, #9 estado#
					415.79, #10 impuesto semestre 1#
					403.79, #11 impuesto semestre 2#
					391.79, #12 rezago#
					367.71, #13 recargo#
					355.79, #14 gto ejecucuion#
					343.79, #15 multa#
					379.79, #16 actualizacion#
					331.751, #17 adeudo predial#
					176.206, #18 nada#
					188.206, #19 consumo#
					176.206, #20 alcantarillado#
					164.206, #21 rezago#
					140.206, #22 recargo#
					152.206, #23 actualizacion#
					128.206, #24 otros cargos#
					116.162, #25 adeudo agua#
					70.823, #26 nada#
					70.823, #27 nada#
					70.823, #28 nada#
					70.823, #29 total final#
					70.823, #30 ignorar#
					431.198, #31 captura predial#
					431.198, #32 captura agua#
					70.823, #33 ignorar#
					70.823, #34 adicional#
					70.823, #35 adicional#

					]



	name = data_dict[35]
	first_name = data_dict[34]

	for data in range(10):
		if data_dict[data] is None:
			data_dict[data] = ""

	data_dict[4] = str(data_dict[4])
	data_dict[5] = str(data_dict[5])
	data_dict[7] = str(data_dict[7])



####################################################
#		BLOQUE 1                                   #
####################################################
	text= '<font name= "Times" color= "#05060B" size="7">%s %s %s %s CP %s <br/> %s %s</font>'%( data_dict[3], data_dict[4], data_dict[5], data_dict[6], data_dict[7], data_dict[8], data_dict[9])
	settingParagraphs1(text, can, stringBuffer, 5, 303, count, data_dict, barcode1, 245, 60, 14.02, 220)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( data_dict[35] )
	settingParagraphs2(text, can, stringBuffer, 62, 343, count, data_dict, barcode1, 195.086, 30.251)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( barcode_value )
	settingParagraphs3(text, can, stringBuffer, 92, 205, count, data_dict, barcode1, 195.086, 30.251)

####################################################
#		BLOQUE 2                                   #
####################################################

	text= '<font name= "Times" color= "#05060B" size="7">%s %s %s %s CP %s <br/> %s %s</font>'%( data_dict[3], data_dict[4], data_dict[5], data_dict[6], data_dict[7], data_dict[8], data_dict[9])
	settingParagraphs1_2(text, can, stringBuffer, 272.7, 303, count, data_dict, barcode1, 245, 60, 278.5, 220)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( data_dict[35] )
	settingParagraphs2_2(text, can, stringBuffer, 330, 343, count, data_dict, barcode1, 195.086, 30.251)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( barcode_value )
	settingParagraphs3_2(text, can, stringBuffer, 356, 205, count, data_dict, barcode1, 195.086, 30.251)

####################################################
#		BLOQUE 3                                   #
####################################################

	text= '<font name= "Times" color= "#05060B" size="7">%s %s %s %s CP %s <br/> %s %s</font>'%( data_dict[3], data_dict[4], data_dict[5], data_dict[6], data_dict[7], data_dict[8], data_dict[9])
	settingParagraphs1_3(text, can, stringBuffer, 535, 303, count, data_dict, barcode1, 245, 60, 542.98, 220)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( data_dict[35] )
	settingParagraphs2_3(text, can, stringBuffer, 593, 343, count, data_dict, barcode1, 195.086, 30.251)

	text= '<font name= "Times" color= "#05060B" size="7">%s</font>'%( barcode_value )
	settingParagraphs3_3(text, can, stringBuffer, 620, 205, count, data_dict, barcode1, 195.086, 30.251)



def settingParagraphs1(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey, barx, bary):
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	barcode1.drawOn(can, barx, bary, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs2(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	count = count + 1
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	#barcode1.drawOn(can, 325, 999, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs3(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	style = getSampleStyleSheet()
	width, height = letter
	paragraph = Paragraph(text, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)


def settingParagraphs1_2(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey, barx, bary):
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	barcode1.drawOn(can, barx, bary, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs2_2(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	count = count + 1
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	#barcode1.drawOn(can, 325, 999, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs3_2(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	style = getSampleStyleSheet()
	width, height = letter
	paragraph = Paragraph(text, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)


def settingParagraphs1_3(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey, barx, bary):
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	barcode1.drawOn(can, barx, bary, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs2_3(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	count = count + 1
	style = getSampleStyleSheet()
	width, height = letter
	frame1 = Frame(x, y, framex, framey, showBoundary=0)
	paragraph = [Paragraph(text, style['Normal'])]
	border = KeepInFrame(6*inch, 15*inch, paragraph)
	borderdraw = frame1.addFromList([border], can)
	#barcode1.drawOn(can, 325, 999, mm)
	#barcode1.drawOn(can, 160.1, 980, mm)
#if count == len(data_dict):
	#saveCanvas(can, stringBuffer)

def settingParagraphs3_3(text,  can, stringBuffer, x, y, count, data_dict, barcode1, framex, framey):
	style = getSampleStyleSheet()
	width, height = letter
	paragraph = Paragraph(text, style=style["Normal"])
	paragraph.wrapOn(can, width, height)
	paragraph.drawOn(can, x, y, mm)
	saveCanvas(can, stringBuffer)

def saveCanvas(can, stringBuffer):
	can.save()
	createPDF(stringBuffer)

def createPDF(stringBuffer):
	global pdf_counter
	#print "createPDF"
	stringBuffer.seek(0)
	new_pdf = PdfFileReader(stringBuffer)
	existing_pdf = PdfFileReader(file("formas.pdf", "rb"))
	pages = existing_pdf.getNumPages()
	output = PdfFileWriter()
	page_1 = existing_pdf.getPage(0)
	#page_2 = existing_pdf.getPage(1)
	page_1.mergePage(new_pdf.getPage(0))
	output.addPage(page_1)
	#output.addPage(page_2)
	name = 'formas_%d.pdf' %pdf_counter
	#print "creating %s"%name
	save_name = os.path.join(os.path.expanduser("~"), "Desktop/FORMS", name)
	outputStream = file(save_name, "wb")
	output.write(outputStream)
	outputStream.close()
	pdf_counter = pdf_counter + 1
	#writePage2(cStringIO.StringIO())

if __name__ == "__main__":
	global pdf_counter, pdf_final, c

	c = 0
	pdf_final= 1
	pdf_counter = 1
	load_wb()
