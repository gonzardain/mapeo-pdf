# -*- coding: utf-8 -*-
import cStringIO
from openpyxl import load_workbook
import math, os, checksum
from reportlab.graphics.barcode import code128, code39
from reportlab.lib.units import mm, cm, inch
import ctypes



def load_wb(count):
	data_dict = []
	wb = load_workbook('base1.xlsx', data_only=False)
	sheet_ranges = wb[wb.get_sheet_names()[0]]
	ws = wb.active
	row_count = ws.max_row
	column_count = ws.max_column
	for row in range(2, row_count+1):
		for column in range (1, column_count+1):
			data = ws.cell(column=column, row=row)
			data = data.value
			create_array(data, column, data_dict, count)

def create_array(data, column, data_dict, count):
	data_dict.append(data)
	if column == 36:
		do_stuff(data_dict, count)


def do_stuff(data_dict, count):
    settingBarcode(data_dict, count)
    del data_dict[:]

def settingBarcode(data_dict, count):
	count= count +1
	value_default = "29"
	value_16 = str(data_dict[31].__add__("0"))
	total_value = "{}{}".format(value_default, value_16)
	checkdigit = str(checksum.add_check_digit(total_value))
	#can = canvas.Canvas(stringBuffer, pagesize=letter)
	barcode_value = "CYF{}{}{}".format(value_default, value_16, checkdigit)
	barcode1 = code39.Standard39(barcode_value, stop=1, barHeight=15*mm, barWidth=.215*mm,  checksum=0)
	#barcode2 = code128.Code128("000715u057123",barWidht=.25*cm,barHeight =5*mm)
	settingText(data_dict, barcode_value, count)


def settingText(data_dict, barcode_value, count):
    #chars_to_remove = ['CD', ',', '.', 'MUN', 'VILLAHERMOSA', 'COL', 'FRAC']
    r= ctypes.create_string_buffer("\0").value

    chars_to_remove = ('CD', r), (',', r), ('.', r), ('MUN', r), ('VILLAHERMOSA', r), ('COL', r), ('FRAC', r), (',', r)
    count = count +1
    file = open('CYF17502.txt','a')
    text=[]
    for data in range(10):
	if data_dict[data] is None:
	    data_dict[data]= ""
        if data == 0:
            text.append("CYF17510|")
        elif data == 1:
            text.append(barcode_value)
            text.append("|")
        elif data == 3:
            text.append(data_dict[data].encode('utf-8'))


        elif data == 4:
            text.append(" ")
            try:
                text.append(str(data_dict[data]))
            except:
                text.append(data_dict[data].encode('utf-8'))
            text.append("|")

        elif data == 6:
            string = data_dict[data]
            colonia = reduce(lambda a, kv: a.replace(*kv), chars_to_remove, string)
            text.append(colonia.encode('utf-8'))
            text.append("|")
            text.append("Centro")
            text.append("|")
            text.append("Tabasco")
            text.append("|")


        elif data == 7:
	    data = str(data_dict[data])
            print data
            text.append(data[:-2])
            text.append("|")
            text.append(data[:-2])


            text.append("\n")
            file.writelines(text)
            save_file(file, text)

	print count


def save_file(file, text):

    file.close()



if __name__ == "__main__":
    count = 0
    load_wb(count)
